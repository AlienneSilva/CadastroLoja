from flask import Flask, request, render, redirect, url_for
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

@app.route("/")
def home():
    return render("index.html")


# Função para cálculo
def calc(valorUni, pecas):
    return float(valorUni) * int(pecas)


@app.route("/gravar", methods=["POST"])
def gravar():

    nome = request.form["nome"]
    contato = request.form["contato"]
    pecas = request.form["pecas"]
    produto = request.form["produto"]
    valorUni = request.form["valorUni"]
    canal = request.form["canal"]
    conta = request.form["conta"]

    valorTotal = calc(valorUni, pecas)

    data_registro = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    arquivo_excel = "cadastros.xlsx"

    if not os.path.exists(arquivo_excel):
        workbook = Workbook()
        planilha = workbook.active
        planilha.title = "Cadastros"

        planilha.append([
            "Nome",
            "Responsável",
            "Peças",
            "Produto",
            "Valor Unitário",
            "Canal",
            "Conta",
            "Valor Total",
            "Data Registro"
        ])
    else:
        workbook = load_workbook(arquivo_excel)
        planilha = workbook.active

    planilha.append([
        nome,
        contato,
        pecas,
        produto,
        valorUni,
        canal,
        conta,
        valorTotal,
        data_registro
    ])

    workbook.save(arquivo_excel)

    return redirect(url_for('home'))
if __name__ == "__main__":
    app.run(debug=True)